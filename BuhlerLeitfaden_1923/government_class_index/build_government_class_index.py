# -*- coding: utf-8 -*-
"""Build the Scherzl/Bühler government (Rektion) + verb-class/voice census
from index_Shertsl_Byuler_dopoln_180721.xlsx, joined to Whitney (present class)
and Tolchelnikov-Talmud 2026 (morphoclass ryad/tip/seṭ + pada).

Emits, to OUTDIR:
  government_lexicon.jsonl        root -> class-readings + governed cases
  government_structure.json       the структура catalogue (197 sub-uses)
  verb_class_voice_census.jsonl   per-root 5-source class + voice + Talmud morphoclass
  verb_class_disagreements.csv    the flagged disagreement subset
Usage: python build_index.py <OUTDIR>
"""
import openpyxl, re, json, csv, unicodedata, collections, sys, os
sys.stdout.reconfigure(encoding='utf-8')

# paths resolve relative to this script's location:
#   <GH>/SanskritGrammar/BuhlerLeitfaden_1923/government_class_index/build_government_class_index.py
# requires sibling clones WhitneyRoots and (in-repo) TolchelnikovTalmud_2026.
HERE = os.path.dirname(os.path.abspath(__file__))
SG   = os.path.dirname(os.path.dirname(HERE))                 # .../SanskritGrammar
GH   = os.path.dirname(SG)                                    # .../GitHub
SRC  = os.path.join(HERE, "..", "index_Shertsl_Byuler_dopoln_180721.xlsx")
WHIT = os.path.join(GH, "WhitneyRoots", "crosswalk", "roots.csv")
TAL  = os.path.join(SG, "TolchelnikovTalmud_2026", "data", "whitney_talmud.json")
OUTDIR = sys.argv[1] if len(sys.argv) > 1 else HERE

ROMAN = {'I':1,'II':2,'III':3,'IV':4,'V':5,'VI':6,'VII':7,'VIII':8,'IX':9,'X':10}
# roots the index spells differently from Whitney's lemma
ALIAS = {'grah':'grabh', 'grabh':'grabh'}

def nk(s):
    if not s: return ''
    s = unicodedata.normalize('NFC', str(s)).strip().lower()
    return s.rstrip('0123456789 ').strip()

def keys_for(root, alt):
    ks = []
    for base in [root, alt]:
        if not base: continue
        for variant in re.split(r'[/,]', base):   # split slash/comma spelling variants
            k = nk(variant)
            if k:
                ks.append(k)
                if k in ALIAS: ks.append(ALIAS[k])
    seen=set(); out=[]
    for k in ks:
        if k not in seen: seen.add(k); out.append(k)
    return out

def roman_set(s):
    return set(ROMAN[t] for t in re.findall(r'\b(?:IX|IV|V?I{0,3}|X)\b', str(s or '')) if t in ROMAN)

def arabic_set(s):
    return set(int(x) for x in re.findall(r'\b(10|[1-9])\b', str(s or '')))

def apte_class_set(txt):
    # leading class digit before a pada letter: '5P.', '2Ā', '9U.', '1 U.'
    return set(int(x) for x in re.findall(r'\b(10|[1-9])\s*\.?\s*[PĀAU]', str(txt or '')))

def voices_in(txt):
    v=set()
    for tok in re.findall(r'\b(P|Ā|A|U)\b\.?', str(txt or '')):
        v.add({'A':'Ā'}.get(tok, tok) if tok=='A' and 'Ā' not in str(txt) else tok)
    # simpler: detect the three canonical pada markers
    out=set()
    if re.search(r'\bU\b', str(txt or '')): out.add('U')
    if re.search(r'Ā', str(txt or '')): out.add('Ā')
    if re.search(r'\bP\b', str(txt or '')): out.add('P')
    return out

# ---------------- government parser ----------------
CASE_RE = re.compile(r'\b(nom|acc|instr|dat|abl|gen|loc|voc)([_\-][a-zA-Zа-яё0-9]+)?\b', re.I)
HEAD_RE = re.compile(r'([^\s\[\]/\n]+)\s*\[\s*([0-9?]+)\s*,\s*([^\]]+?)\]')

def parse_pages(seg):
    out=[]
    for m in re.finditer(r'(\D|^)(\d{1,3})', seg):
        if m.group(1) in ('↓','↑'): continue
        out.append(int(m.group(2)))
    return out

def parse_relations(part):
    ms=list(CASE_RE.finditer(part)); rels=[]
    for i,cm in enumerate(ms):
        prev = part[cm.start()-1] if cm.start()>0 else ''
        if prev=='/':                      # replaceability marker on previous rel
            if rels:
                rels[-1]['replaceable_by']=cm.group(1).lower()+('_'+cm.group(2).lstrip('_-') if cm.group(2) else '')
            continue
        end = ms[i+1].start() if i+1<len(ms) else len(part)
        seg = part[cm.end():end]
        list_pages = parse_pages(re.split(r'\(', seg[:40])[0])
        rels.append({'case':cm.group(1).lower(),
                     'subtype':(cm.group(2) or '').lstrip('_-') or None,
                     'pages':parse_pages(seg[:40]),
                     'hidden':'скрыт' in seg[:35],
                     'example_only':(not list_pages) and bool(re.search(r'\(\s*[-–—]\s*\)|\(\d{1,3}\s*[↓↑]', seg[:25])),
                     'replaceable_by':None})
    return rels

def parse_collated(text):
    if not text: return []
    text=str(text); heads=list(HEAD_RE.finditer(text))
    if not heads:
        stem=text.strip().split()[0] if text.strip() else None
        return [{'stem':stem,'class':None,'voice':None,'relations':parse_relations(text)}]
    out=[]
    for i,h in enumerate(heads):
        end=heads[i+1].start() if i+1<len(heads) else len(text)
        out.append({'stem':h.group(1).strip(),'class':h.group(2),'voice':h.group(3).strip(),
                    'relations':parse_relations(text[h.end():end])})
    return out

# ---------------- load reference sources ----------------
whit={}          # key -> {'class':set(union), 'homonyms':int, 'glosses':[...]}
for row in csv.DictReader(open(WHIT,encoding='utf-8')):
    k=nk(row['root_iast'])
    d=whit.setdefault(k,{'class':set(),'homonyms':0,'glosses':[]})
    d['class'] |= set(ROMAN[x] for x in str(row.get('class','')).split('|') if x in ROMAN)
    d['homonyms']+=1
    d['glosses'].append(row.get('gloss_short',''))

tal={}
tj=json.load(open(TAL,encoding='utf-8'))
for e in tj.get('verbal_roots',[]):
    tal.setdefault(nk(e.get('root_iast')),[]).append(e)

# ---------------- Главный ----------------
wb=openpyxl.load_workbook(SRC,data_only=True); ws=wb['Главный']
# workbook's own human-curated flag: blue fill on Бюлер col = "class differs from dictionaries"
wss=openpyxl.load_workbook(SRC,data_only=False)['Главный']
blue_rows=set()
for ri,srow in enumerate(wss.iter_rows(min_row=3),start=3):
    cell=srow[5]
    f=cell.fill
    if f and f.patternType=='solid' and isinstance(f.fgColor.rgb,str) and f.fgColor.rgb=='FF00B0F0':
        blue_rows.add(ri)
lex=[]; census=[]
for ridx, r in enumerate(ws.iter_rows(values_only=True), start=1):
    if ridx < 3 or not r[1]: continue
    root=str(r[1]).strip(); alt=str(r[2]).strip() if r[2] else None
    class_field=str(r[3]).strip() if r[3] else None
    apte=str(r[4]).strip() if r[4] else None
    buhler=str(r[5]).strip() if r[5] else None
    readings=parse_collated(r[7])
    lex.append({'root':root,'root_alt':alt,'class_field':class_field,'apte_conj':apte,
                'buhler':buhler,'scherzl_orig':str(r[6]).strip() if r[6] else None,
                'readings':readings,'notes':str(r[9]).strip() if len(r)>9 and r[9] else None})
    # ---- census: 5-source class + voice ----
    ks=keys_for(root,alt)
    wk=next((k for k in ks if k in whit),None)
    tk=next((k for k in ks if k in tal),None)
    idx_cls=arabic_set(class_field) | set(int(rd['class']) for rd in readings if rd['class'] and rd['class'].isdigit())
    apte_cls=apte_class_set(apte)
    buhler_cls=roman_set(buhler)
    whit_cls=whit[wk]['class'] if wk else set()
    homonyms=whit[wk]['homonyms'] if wk else 0
    tal_entry=tal[tk][0] if tk else None
    tal_cls=roman_set('|'.join(tal_entry.get('class',[]))) if tal_entry else set()
    idx_voice=set(rd['voice'].split('/')[0].strip() for rd in readings if rd['voice'] and rd['voice'][0] in 'PĀU')
    buhler_voice=voices_in(buhler)
    tal_pada={tal_entry['pada']} if (tal_entry and tal_entry.get('pada')) else set()
    def conflict(a,b): return bool(a) and bool(b) and not (a & b)
    indian = idx_cls | apte_cls          # the Indian dhātupāṭha tradition (index+Apte)
    b_w = conflict(buhler_cls, whit_cls)
    i_w = conflict(indian, whit_cls)
    # classify the disagreement
    if homonyms > 1 and (b_w or i_w):
        dtype = 'join_ambiguous'         # multiple Whitney homonyms -> low confidence
    elif i_w and (not buhler_cls or conflict(buhler_cls, whit_cls)):
        dtype = 'western_vs_indian'      # whole Indian tradition (idx+apte[+buhler]) != Whitney/Talmud
    elif b_w and (buhler_cls and not conflict(indian, whit_cls) and indian):
        dtype = 'buhler_idiosyncratic'   # Bühler alone differs; index+Apte agree with Whitney
    elif b_w or i_w:
        dtype = 'other_conflict'
    else:
        dtype = None
    census.append({
        'root':root,'root_alt':alt,'whitney_key':wk,'talmud_key':tk,
        'class_index':sorted(idx_cls),'class_apte':sorted(apte_cls),'class_buhler':sorted(buhler_cls),
        'class_whitney':sorted(whit_cls),'class_talmud':sorted(tal_cls),
        'buhler_vs_whitney':'conflict' if conflict(buhler_cls,whit_cls) else ('agree' if buhler_cls&whit_cls else 'na'),
        'index_vs_whitney':'conflict' if conflict(idx_cls,whit_cls) else ('agree' if idx_cls&whit_cls else 'na'),
        'apte_vs_whitney':'conflict' if conflict(apte_cls,whit_cls) else ('agree' if apte_cls&whit_cls else 'na'),
        'in_whitney':bool(wk),'in_talmud':bool(tk),'whitney_homonyms':homonyms,
        'disagreement_type':dtype,'workbook_class_flag':ridx in blue_rows,
        'talmud_ryad':tal_entry.get('ryad') if tal_entry else None,
        'talmud_tip':tal_entry.get('tip') if tal_entry else None,
        'talmud_set':tal_entry.get('set') if tal_entry else None,
        'voice_index':sorted(idx_voice),'voice_buhler':sorted(buhler_voice),'voice_talmud_pada':sorted(tal_pada),
        'voice_conflict':bool(idx_voice and tal_pada and not (idx_voice & tal_pada)),
    })

# ---------------- структура ----------------
ws2=wb['структура']; srows=list(ws2.iter_rows(values_only=True))
struct=[]; cur=None
for row in srows[1:]:
    a=(row[0] or '').strip() if row[0] else ''
    if a and ' - ' in a and 'падеж' in a and (not row[1]) and (not row[2]):
        cur=a.split(' - ')[0].strip(); continue
    b=(row[1] or '').strip() if row[1] else ''
    if not (a or b): continue
    c=(row[2] or '').strip() if row[2] else ''
    struct.append({'case':cur,'subdivision':a or None,'context':b or None,
                   'status':{'**':'direct','*':'indirect'}.get(c,'none'),
                   'pages':str(row[3]).strip() if row[3] else None,
                   'notes':str(row[4]).strip() if row[4] else None,
                   'governs_with':str(row[5]).strip() if len(row)>5 and row[5] else None,
                   'question':str(row[6]).strip() if len(row)>6 and row[6] else None,
                   'replaceability':str(row[7]).strip() if len(row)>7 and row[7] else None,
                   'examples':str(row[8]).strip() if len(row)>8 and row[8] else None})

# ---------------- write ----------------
os.makedirs(OUTDIR,exist_ok=True)
def wj(name,rows):
    with open(os.path.join(OUTDIR,name),'w',encoding='utf-8') as f:
        for r in rows: f.write(json.dumps(r,ensure_ascii=False)+'\n')
wj('government_lexicon.jsonl',lex)
wj('verb_class_voice_census.jsonl',census)
json.dump(struct,open(os.path.join(OUTDIR,'government_structure.json'),'w',encoding='utf-8'),ensure_ascii=False,indent=1)

dis=[c for c in census if c['buhler_vs_whitney']=='conflict' or c['index_vs_whitney']=='conflict' or c['apte_vs_whitney']=='conflict' or c['voice_conflict']]
with open(os.path.join(OUTDIR,'verb_class_disagreements.csv'),'w',encoding='utf-8',newline='') as f:
    w=csv.writer(f)
    w.writerow(['root','class_index','class_apte','class_buhler','class_whitney','class_talmud',
                'buhler_vs_whitney','index_vs_whitney','apte_vs_whitney',
                'voice_index','voice_buhler','voice_talmud_pada','voice_conflict','talmud_ryad','talmud_tip','talmud_set'])
    for c in dis:
        w.writerow([c['root'],c['class_index'],c['class_apte'],c['class_buhler'],c['class_whitney'],c['class_talmud'],
                    c['buhler_vs_whitney'],c['index_vs_whitney'],c['apte_vs_whitney'],
                    c['voice_index'],c['voice_buhler'],c['voice_talmud_pada'],c['voice_conflict'],
                    c['talmud_ryad'],c['talmud_tip'],c['talmud_set']])

# ---------------- stats to stdout ----------------
gov_rel=sum(len(rd['relations']) for x in lex for rd in x['readings'])
repl=sum(1 for x in lex for rd in x['readings'] for rl in rd['relations'] if rl['replaceable_by'])
print(f"government_lexicon: roots={len(lex)}  gov_relations={gov_rel}  case_alternations={repl}")
print(f"government_structure: sub-uses={len(struct)}")
print(f"census: rows={len(census)}  in_whitney={sum(c['in_whitney'] for c in census)}  in_talmud={sum(c['in_talmud'] for c in census)}")
# dedupe to unique base roots for the disagreement summary (collapse prefixed-stem duplicates)
by_root={}
for c in census:
    by_root.setdefault(nk(c['root']), c)   # first row per base root
uniq=list(by_root.values())
dt=collections.Counter(c['disagreement_type'] for c in uniq if c['disagreement_type'])
print(f"unique base roots={len(uniq)}")
print("disagreement types (unique base roots):")
for k,v in dt.most_common(): print(f"    {k:22s} {v}")
print(f"  not in Whitney (denominative/secondary)={sum(not c['in_whitney'] for c in uniq)}")
print(f"disagreements.csv rows(all)={len(dis)}")
for label,key in [('western_vs_indian','western_vs_indian'),('buhler_idiosyncratic','buhler_idiosyncratic'),('join_ambiguous','join_ambiguous')]:
    print(f"\n=== {label} (unique base roots) ===")
    for c in uniq:
        if c['disagreement_type']==key:
            print(f"  {c['root']:10s} idx={c['class_index']} apte={c['class_apte']} buhler={c['class_buhler']} whit={c['class_whitney']} tal={c['class_talmud']} hom={c['whitney_homonyms']}")
